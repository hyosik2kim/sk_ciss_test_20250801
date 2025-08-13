import copy
import os

from SCARCommon import *

from openpyxl import load_workbook, Workbook, styles, utils
from parse import *


def info(log):
    if log[:1] == '\n':
        print('')
        log = log.replace('\n', '')

    print(f'[{now()}] [SCARSummarizer] {log}')

class SCARSummarizer():

    summarizeCompleted = None

    def __init__(self):
        
        self.runSummarize  = False

        self.summaryInfos = []
        self.reasonInfos  = {}
        self.transactionInfos = {}

        self.strCommands  = {}
        self.strPrevious  = {}

        self.workbook     = None
        self.worksheet    = None
        self.cellColor    = None
        self.fontColor    = None

        self.summaryColumn = ['Charger', 'Plug #', 'Card', 'Extra Data', 'Start Date', 'End Date', 'Total Time', 'Charging Time', 'Total kWh',
                            'TransactionId', 'StoppedReason', 'Command', 'FromStatus', 'ToStatus', 'SoC', 'Finish', 'Error', 'ErrorPLC', 'Detail', 'SequenceName1', 'SequenceName2']

        self.borderThin = styles.Side(border_style='thin', color='000000')
        self.borderThick = styles.Side(border_style='thick', color='000000')

    def run(self):
        while True:
            if self.runSummarize == True:
                self.summarize()
                self.runSummarize = False
            else:
                self.sleep(1)

    def requestSummarize(self):
        self.runSummarize = True

    def summarize(self):
        self.summaryInfos.clear()
        self.reasonInfos.clear()
        self.transactionInfos.clear()

        self.makeWorkbook()

    def makeWorkbook(self):
        print('')

        info(f'makeWorkbook - Detail [{Config().detailPath}]')

        if os.path.exists(Config().detailPath) == False:
            return

        if os.path.isdir(Config().detailPath) == False:
            self.workbook = load_workbook(Config().detailPath)
        else:
            info(f'Invalid File State [{Config().detailPath}]')
            return

        worksheet = None
        for name in self.workbook.sheetnames:
            if name == 'Summary':
                worksheet = self.workbook[name]
                break

        if worksheet == None:
            self.workbook.create_sheet('Summary')

        self.worksheet = self.workbook['Summary']

        # Detail Sheet
        self.makeSummary()

        # Completed
        try:
            self.workbook.save(Config().detailPath)
            if self.summarizeCompleted :
                self.summarizeCompleted(True)
        except PermissionError:
            info('makeWorkbook - PermissionError')
            if self.summarizeCompleted :
                self.summarizeCompleted(False)

    def makeSummary(self):
        for srcName, logInfos in Analyzed().getLogs():
            info(f'makeSummary [{srcName}] [{len(logInfos)}]')

            self.transactionInfos.clear()
            self.reasonInfos.clear()

            for index, logInfo in enumerate(logInfos):
                if type(logInfo) != LogInfo:
                    continue

                # info(f'Raw    [{index}] -                 Finish [{logInfo.connector2.strCodeFinish}] Error [{logInfo.connector2.strCodeError}] PLC [{logInfo.connector2.strCodeErrorPLC}]')

                if logInfo.ocppMessage.strMessageId == 'Authorize':
                    ocppResponse = parse('{date} {response}', logInfo.ocppMessage.strResponse)
                    if ocppResponse != None and ocppResponse['response'] != 'Accepted':
                        summaryInfo = SummaryInfo()
                        summaryInfo.strCharger = srcName
                        summaryInfo.strConnectorId = '1'
                        summaryInfo.strCard = logInfo.ocppMessage.strIdTag
                        summaryInfo.strExtraData = 'Authorization Failure Reason: Card not found'
                        summaryInfo.strCellColor = 'FF9999'
                        summaryInfo.strStartDate = convertTime(logInfo.ocppMessage.strResponseTime)
                        summaryInfo.strEndDate = summaryInfo.strStartDate
                        summaryInfo.strTarget = logInfo.strTarget
                        self.summaryInfos.append(copy.deepcopy(summaryInfo))

                        summaryInfo.strConnectorId = '2'
                        self.summaryInfos.append(copy.deepcopy(summaryInfo))

                elif logInfo.ocppMessage.strMessageId == 'TransactionEvent':
                    if logInfo.ocppMessage.strEventType == 'Started':
                        summaryInfo = SummaryInfo()
                        summaryInfo.strCharger = srcName
                        summaryInfo.strConnectorId = logInfo.ocppMessage.strConnectorId
                        summaryInfo.strStartDate = convertTime(logInfo.strDateTime)
                        summaryInfo.strStartkWh = logInfo.ocppMessage.strMeterValue
                        summaryInfo.strTransactionId = logInfo.ocppMessage.strTransactionId
                        summaryInfo.strTarget = logInfo.strTarget

                        self.transactionInfos[logInfo.ocppMessage.strTransactionId] = copy.deepcopy(summaryInfo)

                        if summaryInfo.strConnectorId in self.reasonInfos:
                            del(self.reasonInfos[summaryInfo.strConnectorId])

                        # reasonInfo = ReasonInfo()
                        # if summaryInfo.strConnectorId in self.reasonInfos:
                        #     reasonInfo = self.reasonInfos[summaryInfo.strConnectorId]
                        # info(f'Started [{index}] - ConnectorId [{summaryInfo.strConnectorId}] Finish [{reasonInfo.strCodeFinish}] Error [{reasonInfo.strCodeError}] PLC [{reasonInfo.strCodeErrorPLC}]')

                    elif logInfo.ocppMessage.strEventType == 'Updated':
                        if logInfo.ocppMessage.strTransactionId not in self.transactionInfos.keys():
                            continue

                        summaryInfo = copy.deepcopy(self.transactionInfos[logInfo.ocppMessage.strTransactionId])
                        if logInfo.ocppMessage.strChargingState == 'Charging':
                            summaryInfo.strChargingDate = convertTime(logInfo.strDateTime)

                        bBad, strResponse = logInfo.ocppMessage.getBadResponse(bSub=True)
                        if bBad == True and len(strResponse) > 0:
                            summaryInfo.strSubReason = strResponse

                        self.transactionInfos[logInfo.ocppMessage.strTransactionId] = copy.deepcopy(summaryInfo)

                    elif logInfo.ocppMessage.strEventType == 'Ended' and logInfo.ocppMessage.bOffline == False:
                        if logInfo.ocppMessage.strTransactionId in self.transactionInfos.keys():
                            summaryInfo = copy.deepcopy(self.transactionInfos[logInfo.ocppMessage.strTransactionId])
                            summaryInfo.strCard = logInfo.ocppMessage.strIdTag
                            summaryInfo.strEndDate = convertTime(logInfo.strDateTime)
                            summaryInfo.strEndkWh = logInfo.ocppMessage.strMeterValue
                            summaryInfo.strStopReason = logInfo.ocppMessage.strStoppedReason

                            for connectorId, reasonInfo in self.reasonInfos.items():
                                if summaryInfo.strConnectorId == connectorId:
                                    summaryInfo.reasonInfo = copy.deepcopy(reasonInfo)
                                    break

                            # if summaryInfo.strConnectorId in self.reasonInfos:
                            #     reasonInfo = self.reasonInfos[summaryInfo.strConnectorId]
                            # info(f'Ended  [{index}] - ConnectorId [{summaryInfo.strConnectorId}] Finish [{reasonInfo.strCodeFinish}] Error [{reasonInfo.strCodeError}] PLC [{reasonInfo.strCodeErrorPLC}]')
                            # print('')

                            self.transactionInfos[logInfo.ocppMessage.strTransactionId] = copy.deepcopy(summaryInfo)

                            if logInfo.ocppMessage.strStoppedReason == 'StoppedByEV':
                                summaryInfo.strExtraData = 'Stop Reason: The transaction was stopped by the EV'
                            elif logInfo.ocppMessage.strStoppedReason == 'DeAuthorized':
                                if summaryInfo.strSubReason == 'Blocked':   # Authorize [Accepted] > TransactionEvent [Blocked]
                                    summaryInfo.strExtraData = 'Authorization Failure Reason: Rejected by billing service'
                                else:
                                    summaryInfo.strExtraData = 'Stop Reason: Authorization removed'
                            elif logInfo.ocppMessage.strStoppedReason == 'Local':
                                summaryInfo.strExtraData = 'Stop Reason: Driver stopped'
                            elif logInfo.ocppMessage.strStoppedReason == 'Timeout':
                                summaryInfo.strExtraData = 'Stop Reason: EV not connected within timeout'
                            elif logInfo.ocppMessage.strStoppedReason == 'Other':
                                summaryInfo.strExtraData = 'Stop Reason: Other'
                            elif logInfo.ocppMessage.strStoppedReason == 'SOCLimitReached':
                                summaryInfo.strExtraData = 'Stop Reason: Electric vehicle has reported reaching a locally enforced maximum battery State of Char...'
                            elif logInfo.ocppMessage.strStoppedReason == 'PowerLoss':
                                summaryInfo.strExtraData = 'Stop Reason: Power loss'
                            elif logInfo.ocppMessage.strStoppedReason == 'EmergencyStop':
                                summaryInfo.strExtraData = 'Stop Reason: Emergency stop'
                            elif logInfo.ocppMessage.strStoppedReason == 'Remote':
                                summaryInfo.strExtraData = 'Stop Reason: Server stopped'
                            elif logInfo.ocppMessage.strStoppedReason == 'EVDisconnected':
                                summaryInfo.strExtraData = 'Stop Reason: EV disconnected'

                            # elif logInfo.ocppMessage.strStoppedReason == '':
                            #     summaryInfo.strExtraData = 'Stop Reason: Manual stop charge by operator'

                            summaryInfo.reasonInfo.strSoC = logInfo.ocppMessage.strSoC

                            if summaryInfo.getChargingTime() >= 300:
                                summaryInfo.strCellColor = 'A6C9EC'

                            self.summaryInfos.append(summaryInfo)

                            if logInfo.ocppMessage.strTransactionId in self.transactionInfos:
                                del(self.transactionInfos[logInfo.ocppMessage.strTransactionId])
                            if summaryInfo.strConnectorId in self.reasonInfos:
                                del(self.reasonInfos[summaryInfo.strConnectorId])

                elif len(logInfo.strCommand) > 0 and logInfo.strCommand != "Init":
                    for strConnector in ['1', '2']:
                        if strConnector not in self.reasonInfos:
                            self.strCommands[strConnector] = ''

                        self.strCommands[strConnector] = logInfo.strCommand

                elif len(logInfo.connector1.strPreviousStatus) > 0:
                    if '1' not in self.reasonInfos:
                        self.strPrevious['1'] = ''

                    self.strPrevious['1'] = logInfo.connector1.strPreviousStatus

                elif len(logInfo.connector1.strCurrentStatus) > 0:
                    if '1' not in self.reasonInfos:
                        self.reasonInfos['1'] = ReasonInfo()

                    self.reasonInfos['1'].strCommand        = self.strCommands['1'] if '1' in self.strCommands else ''
                    self.reasonInfos['1'].strPreviousStatus = self.strPrevious['1'] if '1' in self.strPrevious else ''
                    self.reasonInfos['1'].strCurrentStatus  = logInfo.connector1.strCurrentStatus

                    if logInfo.connector1.strCurrentStatus == 'ready':
                        self.reasonInfos['1'].strSequenceName1 = ''
                        self.reasonInfos['1'].strSequenceName2 = ''

                elif logInfo.connector1.isValid() == True:
                    if '1' not in self.reasonInfos:
                        self.reasonInfos['1'] = ReasonInfo()

                    self.reasonInfos['1'].strCodeFinish   = logInfo.connector1.strCodeFinish
                    self.reasonInfos['1'].strCodeError    = logInfo.connector1.strCodeError
                    self.reasonInfos['1'].strCodeErrorPLC = logInfo.connector1.strCodeErrorPLC

                    # info(f'Reason [{index}] - ConnectorId [1] Finish [{logInfo.connector1.strCodeFinish}] Error [{logInfo.connector1.strCodeError}] PLC [{logInfo.connector1.strCodeErrorPLC}]')

                elif len(logInfo.connector1.strSequenceName1) > 0 and len(logInfo.connector1.strSequenceName2) > 0:
                    if '1' not in self.reasonInfos:
                        self.reasonInfos['1'] = ReasonInfo()

                    self.reasonInfos['1'].strSequenceName1 = logInfo.connector1.strSequenceName1
                    self.reasonInfos['1'].strSequenceName2 = logInfo.connector1.strSequenceName2

                elif len(logInfo.connector2.strPreviousStatus) > 0:
                    if '2' not in self.reasonInfos:
                        self.strPrevious['2'] = ''

                    self.strPrevious['2'] = logInfo.connector2.strPreviousStatus

                elif len(logInfo.connector2.strCurrentStatus) > 0:
                    if '2' not in self.reasonInfos:
                        self.reasonInfos['2'] = ReasonInfo()

                    self.reasonInfos['2'].strCommand        = self.strCommands['2'] if '2' in self.strCommands else ''
                    self.reasonInfos['2'].strPreviousStatus = self.strPrevious['2'] if '2' in self.strPrevious else ''
                    self.reasonInfos['2'].strCurrentStatus  = logInfo.connector2.strCurrentStatus

                    if logInfo.connector2.strCurrentStatus == 'ready':
                        self.reasonInfos['2'].strSequenceName1 = ''
                        self.reasonInfos['2'].strSequenceName2 = ''

                elif logInfo.connector2.isValid() == True:
                    if '2' not in self.reasonInfos:
                        self.reasonInfos['2'] = ReasonInfo()

                    self.reasonInfos['2'].strCodeFinish   = logInfo.connector2.strCodeFinish
                    self.reasonInfos['2'].strCodeError    = logInfo.connector2.strCodeError
                    self.reasonInfos['2'].strCodeErrorPLC = logInfo.connector2.strCodeErrorPLC

                    # info(f'Reason [{index}] - ConnectorId [2] Finish [{logInfo.connector2.strCodeFinish}] Error [{logInfo.connector2.strCodeError}] PLC [{logInfo.connector2.strCodeErrorPLC}]')

                elif len(logInfo.connector2.strSequenceName1) > 0 and len(logInfo.connector2.strSequenceName2) > 0:
                    if '2' not in self.reasonInfos:
                        self.reasonInfos['2'] = ReasonInfo()

                    self.reasonInfos['2'].strSequenceName1 = logInfo.connector2.strSequenceName1
                    self.reasonInfos['2'].strSequenceName2 = logInfo.connector2.strSequenceName2

        self.fontColor = None

        self.worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=10)
        self.setWorkcell(2, 3, f'Server', color='D8D8D8')

        self.worksheet.cell(2, 11).border = styles.Border(left=self.borderThin)
        self.worksheet.merge_cells(start_row=2, start_column=11, end_row=2, end_column=20)
        self.setWorkcell(2, 11, f'System Log', color='D8D8D8')

        self.worksheet.cell(2, 21).border = styles.Border(left=self.borderThin)
        self.worksheet.merge_cells(start_row=2, start_column=21, end_row=2, end_column=22)  # end_column=50
        self.setWorkcell(2, 21, f'Sequence Info', color='D8D8D8')

        index = 0
        for column in self.summaryColumn:
            self.setWorkcell(3, 2 + index, column, color='D8D8D8')

            border = None
            if index == 9 or index == 19:
                border = self.borderThin
            self.worksheet.cell(3, 2 + index).border = styles.Border(left=border, bottom=self.borderThick)

            index += 1

        index = 0
        for summaryInfo in sorted(self.summaryInfos, key=lambda x: getattr(x, 'strStartDate'), reverse=False):
            self.cellColor = None
            if index % 2 == 1:
                self.cellColor = 'F2F2F2'

            if len(summaryInfo.strCellColor) > 0:
                self.cellColor = summaryInfo.strCellColor

            self.setWorkcell(4 + index, 2, summaryInfo.strCharger)                      # 'Charger'
            self.setWorkcell(4 + index, 3, summaryInfo.strConnectorId)                  # 'Plug'
            self.setWorkcell(4 + index, 4, summaryInfo.strCard)                         # 'Card'
            self.setWorkcell(4 + index, 5, summaryInfo.strExtraData, link=summaryInfo.getHyperLink())   # 'Extra Data'
            self.setWorkcell(4 + index, 6, summaryInfo.strStartDate)                    # 'Start Date'
            self.setWorkcell(4 + index, 7, summaryInfo.strEndDate)                      # 'End Date'
            self.setWorkcell(4 + index, 8, secsToStr(summaryInfo.getTotalTime()))       # 'Total Time'
            self.setWorkcell(4 + index, 9, secsToStr(summaryInfo.getChargingTime()))    # 'Charging Time'
            self.setWorkcell(4 + index, 10, summaryInfo.getTotalkWh())                  # 'Total kWh'
            self.setWorkcell(4 + index, 11, summaryInfo.strTransactionId)               # 'TransactionId'
            self.setWorkcell(4 + index, 12, summaryInfo.strStopReason)                  # 'StopReason'
            self.setWorkcell(4 + index, 13, summaryInfo.reasonInfo.strCommand)          # 'Command'
            self.setWorkcell(4 + index, 14, summaryInfo.reasonInfo.strPreviousStatus)   # 'FromStatus'
            self.setWorkcell(4 + index, 15, summaryInfo.reasonInfo.strCurrentStatus)    # 'ToStatus'
            self.setWorkcell(4 + index, 16, summaryInfo.reasonInfo.strSoC)              # 'SoC'
            self.setWorkcell(4 + index, 17, summaryInfo.reasonInfo.getCodeFinish())     # 'Finish'
            self.setWorkcell(4 + index, 18, summaryInfo.reasonInfo.getCodeError())      # 'Error'
            self.setWorkcell(4 + index, 19, summaryInfo.reasonInfo.getCodeErrorPLC())   # 'ErrorPLC'
            self.setWorkcell(4 + index, 20, summaryInfo.reasonInfo.getDetailCode())     # 'DetailCode'
            self.setWorkcell(4 + index, 21, summaryInfo.reasonInfo.getSequenceName1())  # 'SequenceName1'
            self.setWorkcell(4 + index, 22, summaryInfo.reasonInfo.getSequenceName2())  # 'SequenceName2'

            # for sequenceIdx, sequenceData in enumerate(summaryInfo.reasonInfo.getSequenceDetail()):
            #     if sequenceIdx >= 29:
            #         break
            #     self.setWorkcell(4 + index, 23 + sequenceIdx, sequenceData)    # 'Sequence'

            self.worksheet.cell(4 + index, 10).border = styles.Border(right=self.borderThin)
            self.worksheet.cell(4 + index, 20).border = styles.Border(right=self.borderThin)

            index += 1

        self.worksheet.auto_filter.ref = f'B3:R{3 + index}'
        self.worksheet.freeze_panes = 'C4'
        self.worksheet.sheet_view.zoomScale = 75
        for column_cells in self.worksheet.columns:
            length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
            self.worksheet.column_dimensions[column_cells[0].column_letter].width = length

        self.worksheet.column_dimensions['D'].width = 20
        self.worksheet.column_dimensions['E'].width = 60

    def setWorkcell(self, row, column, text, center=True, color=None, red=False, link=None):
        if text != None and len(str(text)) > 0:
            if link != None:
                self.worksheet.cell(row, column, f'=HYPERLINK("{link}", "{text}")')
            else:
                self.worksheet.cell(row, column, text)

        if center == True:
            self.worksheet.cell(row, column).alignment = styles.Alignment(horizontal='center')

        if red == True:
            self.worksheet.cell(row, column).font = styles.Font(name='맑은 고딕', size=11, bold=True, color='FF0000')
        elif self.fontColor != None:
            self.worksheet.cell(row, column).font = styles.Font(name='맑은 고딕', size=11, color=self.fontColor)

        background = self.cellColor
        if color != None:
            background = color
        if background != None:
            self.worksheet.cell(row, column).fill = styles.PatternFill(start_color=background, end_color=background, fill_type='solid')
