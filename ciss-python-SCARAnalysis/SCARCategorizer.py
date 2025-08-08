import copy
import os

from SCARCommon import *

from openpyxl import load_workbook, Workbook, styles, utils
from parse import *

def info(log):
    if log[:1] == '\n':
        print('')
        log = log.replace('\n', '')

    print(f'[{now()}] [SCARCategorizer] {log}')

class SCARCategorizer :

    categorizeCompleted = None

    def __init__(self):
        
        self.runSave      = False
        self.runLoad      = False

        self.workbook     = None
        self.worksheet    = None
        self.cellColor    = None
        self.fontColor    = None

        self.detailColumn = ['DateTime', 'EVSE Event', 'Command', 'SequenceName1', 'SequenceName2',
                            'Connection', 'Status', 'Noti', 'Finish', 'Error', 'ErrorPLC',
                            'Connection', 'Status', 'Noti', 'Finish', 'Error', 'ErrorPLC',
                            'MessageId', 'Response', 'UUID', 'ConnectorId', 'EventType', 'ChargingState', 'TriggerReason', 'StoppedReason', 'TransactionId', 'IdTag', 'MeterValue', 'SoC']

        self.borderThin = styles.Side(border_style='thin', color='000000')
        self.borderThick = styles.Side(border_style='thick', color='000000')

    def run(self):
        while True:
            if self.runSave == True:
                self.makeWorkbook()
                self.runSave = False
            elif self.runLoad == True:
                self.readWorkbook()
                self.runLoad = False
            else:
                self.sleep(1)

    def saveCategorize(self):
        self.runSave = True

    def loadCategorize(self):
        self.runLoad = True

    def makeWorkbook(self):
        print('')

        if os.path.exists(Config().detailPath) == False:
            self.workbook = Workbook()
        else:
            if os.path.isdir(Config().detailPath) == False:
                self.workbook = load_workbook(Config().detailPath)
            else:
                info(f'Invalid File State [{Config().detailPath}]')
                return

        worksheet = None
        for name in self.workbook.sheetnames:
            if name == 'Sheet':
                worksheet = self.workbook[name]
                worksheet.title = 'Summary'
            else:
                self.workbook.remove(self.workbook[name])

        if worksheet == None:
            self.workbook.create_sheet('Summary')

        # Detail Sheet
        if self.makeDetail() == True:
            # Completed
            try:
                self.workbook.save(Config().detailPath)
                if self.categorizeCompleted :
                    self.categorizeCompleted(True)
            except PermissionError:
                info('makeWorkbook - PermissionError')
                if self.categorizeCompleted :
                    self.categorizeCompleted(False)

    def makeDetail(self):
        for srcName, logInfos in Analyzed().getLogs():
            info(f'makeDetail [{srcName}] [{len(logInfos)}]')

            self.fontColor = None
            worksheet = None
            for name in self.workbook.sheetnames:
                if name == 'Sheet':
                    worksheet = self.workbook[name]
                    worksheet.title = srcName
                    break
                elif name == srcName:
                    self.workbook.remove(self.workbook[srcName])
                    self.workbook.create_sheet(srcName)
                    worksheet = self.workbook[srcName]
                    break

            if worksheet == None:
                self.workbook.create_sheet(srcName)

            try:
                self.workbook.save(Config().detailPath)
            except PermissionError:
                info('makeWorkbook - PermissionError')
                if self.categorizeCompleted : 
                    self.categorizeCompleted(False)
                return False

            self.worksheet = self.workbook[srcName]

            self.worksheet.cell(2, 5).border = styles.Border(right=self.borderThin)
            self.worksheet.merge_cells(start_row=2, start_column=6, end_row=2, end_column=11)
            self.setWorkcell(2, 6, f'Connector #1', color='D8D8D8')

            self.worksheet.cell(2, 11).border = styles.Border(right=self.borderThin)
            self.worksheet.merge_cells(start_row=2, start_column=12, end_row=2, end_column=17)
            self.setWorkcell(2, 12, f'Connector #2', color='D8D8D8')

            self.worksheet.cell(2, 17).border = styles.Border(right=self.borderThin)
            self.worksheet.merge_cells(start_row=2, start_column=18, end_row=2, end_column=29)
            self.setWorkcell(2, 18, f'OCPP Message', color='D8D8D8')

            index = 0
            for column in self.detailColumn:
                self.setWorkcell(3, 1 + index, column, color='D8D8D8')

                border = None
                if index == 4 or index == 10 or index == 16:
                    border = self.borderThin
                self.worksheet.cell(3, 1 + index).border = styles.Border(right=border, bottom=self.borderThick)

                index += 1

            index = 0
            for logIdx, logInfo in enumerate(logInfos):
                if type(logInfo) != LogInfo:
                    continue
                elif len(logInfo.connector1.strPreviousStatus) > 0 or len(logInfo.connector2.strPreviousStatus) > 0:
                    continue

                self.cellColor = None
                if index % 2 == 1:
                    self.cellColor = 'F2F2F2'

                self.fontColor = None
                if logInfo.ocppMessage.bOffline == True:
                    self.fontColor = '909090'

                self.setWorkcell(4 + index, 1, logInfo.strDateTime)                     # 'DateTime'
                self.setWorkcell(4 + index, 2, logInfo.strEVSEEvent)                    # 'EVSEEvent'
                self.setWorkcell(4 + index, 3, logInfo.strCommand)                      # 'Command'
                self.setWorkcell(4 + index, 4, logInfo.connector1.strSequenceName1 if len(logInfo.connector1.strSequenceName1) > 0 else logInfo.connector2.strSequenceName1)    # 'SequenceName1'
                self.setWorkcell(4 + index, 5, logInfo.connector1.strSequenceName2 if len(logInfo.connector1.strSequenceName1) > 0 else logInfo.connector2.strSequenceName2)    # 'SequenceName2'

                self.setWorkcell(4 + index, 6, logInfo.connector1.getConnection())      # 'Connection'
                self.setWorkcell(4 + index, 7, logInfo.connector1.strCurrentStatus if len(logInfo.connector1.strCurrentStatus) > 0 else logInfo.connector1.strPreviousStatus)     # 'Status'
                self.setWorkcell(4 + index, 8, logInfo.connector1.strNotification, red=logInfo.connector1.getBadStatus())  # 'Noti'
                self.setWorkcell(4 + index, 9, logInfo.connector1.getCodeFinish())      # 'Finish'
                self.setWorkcell(4 + index, 10, logInfo.connector1.getCodeError())      # 'Error'
                self.setWorkcell(4 + index, 11, logInfo.connector1.getCodeErrorPLC())   # 'ErrorPLC'

                self.setWorkcell(4 + index, 12, logInfo.connector2.getConnection())     # 'Connection'
                self.setWorkcell(4 + index, 13, logInfo.connector2.strCurrentStatus if len(logInfo.connector2.strCurrentStatus) > 0 else logInfo.connector2.strPreviousStatus)    # 'Status'
                self.setWorkcell(4 + index, 14, logInfo.connector2.strNotification, red=logInfo.connector2.getBadStatus()) # 'Noti'
                self.setWorkcell(4 + index, 15, logInfo.connector2.getCodeFinish())     # 'Finish'
                self.setWorkcell(4 + index, 16, logInfo.connector2.getCodeError())      # 'Error'
                self.setWorkcell(4 + index, 17, logInfo.connector2.getCodeErrorPLC())   # 'ErrorPLC'

                self.setWorkcell(4 + index, 18, logInfo.ocppMessage.strMessageId)       # 'MessageId'
                self.setWorkcell(4 + index, 19, logInfo.ocppMessage.strResponse, red=logInfo.ocppMessage.getBadResponse()) # 'Response'
                self.setWorkcell(4 + index, 20, logInfo.ocppMessage.strUUID)            # 'UUID'
                self.setWorkcell(4 + index, 21, logInfo.ocppMessage.strConnectorId)     # 'ConnectorId'
                self.setWorkcell(4 + index, 22, logInfo.ocppMessage.strEventType)       # 'EventType'
                self.setWorkcell(4 + index, 23, logInfo.ocppMessage.strChargingState)   # 'ChargingState'
                self.setWorkcell(4 + index, 24, logInfo.ocppMessage.strTriggerReason)   # 'TriggerReason'
                self.setWorkcell(4 + index, 25, logInfo.ocppMessage.strStoppedReason)   # 'StoppedReason'
                self.setWorkcell(4 + index, 26, logInfo.ocppMessage.strTransactionId)   # 'TransactionId'
                self.setWorkcell(4 + index, 27, logInfo.ocppMessage.strIdTag)           # 'IdTag'
                self.setWorkcell(4 + index, 28, logInfo.ocppMessage.strMeterValue)      # 'MeterValue'
                self.setWorkcell(4 + index, 29, logInfo.ocppMessage.strSoC)             # 'SoC'

                self.worksheet.cell(4 + index, 5).border = styles.Border(right=self.borderThin)
                self.worksheet.cell(4 + index, 11).border = styles.Border(right=self.borderThin)
                self.worksheet.cell(4 + index, 17).border = styles.Border(right=self.borderThin)

                if logInfo.ocppMessage.strMessageId != 'TransactionEvent' or logInfo.ocppMessage.strEventType == 'Started':
                    logInfo.strTarget = f'{utils.get_column_letter(18)}{4 + index}'

                if len(logInfo.strTarget) > 0:
                    Analyzed().setLogs(srcName, logIdx, logInfo)

                index += 1

            self.worksheet.freeze_panes = 'B4'
            self.worksheet.sheet_view.zoomScale = 55
            for column_cells in self.worksheet.columns:
                length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
                self.worksheet.column_dimensions[column_cells[0].column_letter].width = length

        return True

    def readWorkbook(self):
        print('')

        if os.path.exists(Config().detailPath) == False:
            return

        if os.path.isdir(Config().detailPath) == False:
            self.workbook = load_workbook(Config().detailPath)
        else:
            info(f'Invalid File State [{Config().detailPath}]')
            return

        # info('readWorkbook')

        self.worksheet = None
        for name in self.workbook.sheetnames:
            if name != 'Summary':
                self.worksheet = self.workbook[name]
                self.readDetail()
        if self.categorizeCompleted :
            self.categorizeCompleted(self.worksheet != None)

    def readDetail(self):
        rowIndex = 4
        strPreviousStatus = { 1:'None' , 2:'None' }
        strPreviousSequence = { 1:'' , 2:'' }
        while True:
            firstCell = self.worksheet.cell(row=rowIndex, column=1).value
            if firstCell is None:
                break  # 첫 번째 열이 비어 있으면 종료

            rowData = [self.worksheet.cell(row=rowIndex, column=colIndex).value for colIndex in range(1, 30)]
            if len(rowData) > 0:
                logInfo = LogInfo()
                logInfo.clear()
                logInfo.strDateTime         = rowData[0] if rowData[0] != None else ''
                logInfo.strEVSEEvent        = rowData[1] if rowData[1] != None else ''
                logInfo.strCommand          = rowData[2] if rowData[2] != None else ''

                connector = ConnectorInfo()
                connector.clear()
                connector.nConnection       = 1 if rowData[5] == 'On' else (2 if rowData[5] == 'Off' else 0)
                connector.strCurrentStatus  = rowData[6] if rowData[6] != None else ''
                connector.strNotification   = rowData[7] if rowData[7] != None else ''
                connector.strCodeFinish     = rowData[8] if rowData[8] != None else ''
                connector.strCodeError      = rowData[9] if rowData[9] != None else ''
                connector.strCodeErrorPLC   = rowData[10] if rowData[10] != None else ''
                if strPreviousStatus[1] == 'finish' or strPreviousStatus[1] == 'thankYou' or strPreviousStatus[1] == 'fault':
                    if len(strPreviousSequence[1]) <= 0:
                        connector.strSequenceName1  = rowData[3] if rowData[3] != None else ''
                        connector.strSequenceName2  = rowData[4] if rowData[4] != None else ''
                        strPreviousSequence[1] = f'{connector.strSequenceName1}{connector.strSequenceName2}'
                else:
                    strPreviousSequence[1] = ''
                logInfo.connector1 = copy.deepcopy(connector)

                connector.clear()
                connector.nConnection       = 1 if rowData[11] == 'On' else (2 if rowData[11] == 'Off' else 0)
                connector.strCurrentStatus  = rowData[12] if rowData[12] != None else ''
                connector.strNotification   = rowData[13] if rowData[13] != None else ''
                connector.strCodeFinish     = rowData[14] if rowData[14] != None else ''
                connector.strCodeError      = rowData[15] if rowData[15] != None else ''
                connector.strCodeErrorPLC   = rowData[16] if rowData[16] != None else ''
                if strPreviousStatus[2] == 'finish' or strPreviousStatus[2] == 'thankYou' or strPreviousStatus[2] == 'fault':
                    if len(strPreviousSequence[2]) <= 0:
                        connector.strSequenceName1  = rowData[3] if rowData[3] != None else ''
                        connector.strSequenceName2  = rowData[4] if rowData[4] != None else ''
                        strPreviousSequence[2] = f'{connector.strSequenceName1}{connector.strSequenceName2}'
                else:
                    strPreviousSequence[2] = ''
                logInfo.connector2 = copy.deepcopy(connector)

                ocppMessage = OCPPMessageInfo()
                ocppMessage.clear()
                ocppMessage.strMessageId     = rowData[17] if rowData[17] != None else ''
                ocppMessage.strResponse      = rowData[18] if rowData[18] != None else ''
                ocppMessage.strUUID          = rowData[19] if rowData[19] != None else ''
                ocppMessage.strConnectorId   = rowData[20] if rowData[20] != None else ''
                ocppMessage.strEventType     = rowData[21] if rowData[21] != None else ''
                ocppMessage.strChargingState = rowData[22] if rowData[22] != None else ''
                ocppMessage.strTriggerReason = rowData[23] if rowData[23] != None else ''
                ocppMessage.strStoppedReason = rowData[24] if rowData[24] != None else ''
                ocppMessage.strTransactionId = rowData[25] if rowData[25] != None else ''
                ocppMessage.strIdTag         = rowData[26] if rowData[26] != None else ''
                ocppMessage.strMeterValue    = rowData[27] if rowData[27] != None else ''
                ocppMessage.strSoC           = rowData[28] if rowData[28] != None else ''
                if ocppMessage.strMessageId == 'Authorize':
                    strTime     = ocppMessage.strResponse
                    strDateTime = logInfo.strDateTime.split(' ', maxsplit=1)
                    strOffset   = strDateTime[1].split('-', maxsplit=1)[1]
                    if ocppMessage.strResponse.find(' ') > 0:
                        strTime = ocppMessage.strResponse.split(' ', maxsplit=1)[0]

                    ocppMessage.strResponseTime = f'{strDateTime[0]} {strTime}-{strOffset}'

                logInfo.ocppMessage = copy.deepcopy(ocppMessage)

                if logInfo.strEVSEEvent == 'Start App':
                    strPreviousStatus = {1:'None' , 2:'None'}

                elif len(logInfo.connector1.strCurrentStatus) > 0:
                    statusInfo = LogInfo()
                    statusInfo.connector1.strPreviousStatus = strPreviousStatus[1]
                    Analyzed().append(self.worksheet.title, statusInfo)
                    strPreviousStatus[1] = logInfo.connector1.strCurrentStatus

                elif len(logInfo.connector2.strCurrentStatus) > 0:
                    statusInfo = LogInfo()
                    statusInfo.connector2.strPreviousStatus = strPreviousStatus[2]
                    Analyzed().append(self.worksheet.title, statusInfo)
                    strPreviousStatus[2] = logInfo.connector2.strCurrentStatus

                if logInfo.ocppMessage.strMessageId != 'TransactionEvent' or logInfo.ocppMessage.strEventType == 'Started':
                    logInfo.strTarget = f'{utils.get_column_letter(18)}{rowIndex}'

                Analyzed().append(self.worksheet.title, logInfo)

            rowIndex += 1

        info(f'readDetail - Src [{self.worksheet.title}] Size [{Analyzed().sizeLogs(self.worksheet.title)}]')

    def setWorkcell(self, row, column, text, center=True, color=None, red=False, link=None):
        if text != None and len(str(text)) > 0:
            if link != None:
                self.worksheet.cell(row, column, f'=HYPERLINK("{link}", "{text}")')
            else:
                self.worksheet.cell(row, column, text)

        if center == True:
            self.worksheet.cell(row, column).alignment = styles.Alignment(horizontal='center')

        if red == True:
            self.worksheet.cell(row, column).font = styles.Font(name='맑은 고딕', size=11, bold=True, color='FF0000')
        elif self.fontColor != None:
            self.worksheet.cell(row, column).font = styles.Font(name='맑은 고딕', size=11, color=self.fontColor)

        background = self.cellColor
        if color != None:
            background = color
        if background != None:
            self.worksheet.cell(row, column).fill = styles.PatternFill(start_color=background, end_color=background, fill_type='solid')
