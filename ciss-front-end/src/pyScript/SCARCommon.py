import copy
import json
import os
import threading
import yaml
import zipfile

from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from parse import *
from typing import Dict

class LogElement(object):
    TIME    = 0
    MESSAGE = 1
    ORIGIN  = 2

class Config:
    _instance = None

    def __new__(cls):
        if cls._instance == None:
            cls._instance = super(Config, cls).__new__(cls)
            cls._instance.logPath    = ''
            cls._instance.logName    = ''
            cls._instance.logExt     = ''
            cls._instance.zipExt     = ''
            cls._instance.dstFile    = ''
            cls._instance.detailFile = ''
            cls._instance.detailPath = ''
            cls._instance.siteFile   = ''
            cls._instance.siteList   = {}
        return cls._instance

    def clear(self):
        self.logPath    = ''
        self.logName    = ''
        self.logExt     = ''
        self.zipExt     = ''
        self.dstFile    = ''
        self.detailFile = ''
        self.detailPath = ''
        self.siteFile   = ''
        self.siteList   = {}

class LogList:
    _instance = None

    def __new__(cls):
        if cls._instance == None:
            cls._instance = super(LogList, cls).__new__(cls)
            cls._instance.srcDicts  = {}
            cls._instance.dstDicts  = {}
            cls._instance.fileCount = 10
        return cls._instance

    def clear(self):
        self.srcDicts  = {}
        self.dstDicts  = {}
        self.fileCount = 10

class ConnectorInfo:
    def __init__(self):
        self.nConnection       = 0
        self.strPreviousStatus = ''
        self.strCurrentStatus  = ''
        self.strNotification   = ''
        self.strCodeFinish     = ''
        self.strCodeError      = ''
        self.strCodeErrorPLC   = ''
        self.strSequenceName1  = ''
        self.strSequenceName2  = ''

    def clear(self):
        self.nConnection       = 0
        self.strPreviousStatus = ''
        self.strCurrentStatus  = ''
        self.strNotification   = ''
        self.strCodeFinish     = ''
        self.strCodeError      = ''
        self.strCodeErrorPLC   = ''
        self.strSequenceName1  = ''
        self.strSequenceName2  = ''

    def getConnection(self):
        strConnection = ''
        if self.nConnection == 1:
            strConnection = 'On'
        elif self.nConnection == 2:
            strConnection = 'Off'

        return strConnection

    def getCodeFinish(self):
        return self.strCodeFinish if len(self.strCodeFinish) > 0 else ''

    def getCodeError(self):
        return self.strCodeError if len(self.strCodeError) > 0 else ''

    def getCodeErrorPLC(self):
        return self.strCodeErrorPLC if len(self.strCodeErrorPLC) > 0 else ''

    def getBadStatus(self):
        return (len(self.strNotification) > 0 and (self.strNotification == 'Faulted' or self.strNotification == 'Unavailable'))

    def isValid(self):
        return (len(self.strCodeFinish) > 0 and self.strCodeFinish != '0') or \
               (len(self.strCodeError) > 0 and self.strCodeError != '0') or \
               (len(self.strCodeErrorPLC) > 0 and self.strCodeErrorPLC != '0')

class OCPPMessageInfo:
    def __init__(self):
        self.strMessageId     = ''
        self.strResponse      = ''
        self.strResponseTime  = ''
        self.strUUID          = ''
        self.strConnectorId   = ''
        self.strEventType     = ''
        self.strChargingState = ''
        self.strStoppedReason = ''
        self.strTriggerReason = ''
        self.strTransactionId = ''
        self.strIdTag         = ''
        self.strSoC           = ''
        self.strMeterValue    = ''
        self.strTimestamp     = ''
        self.bOffline         = False

    def clear(self):
        self.strMessageId     = ''
        self.strResponse      = ''
        self.strResponseTime  = ''
        self.strUUID          = ''
        self.strConnectorId   = ''
        self.strEventType     = ''
        self.strChargingState = ''
        self.strStoppedReason = ''
        self.strTriggerReason = ''
        self.strTransactionId = ''
        self.strIdTag         = ''
        self.strSoC           = ''
        self.strMeterValue    = ''
        self.strTimestamp     = ''
        self.bOffline         = False

    def getBadResponse(self, bSub=False):
        ocppResponse = parse('{time} {response}', self.strResponse)
        bResponse = False
        strResponse = ''

        if ocppResponse != None and len(ocppResponse['response']) > 0:
            bResponse = bool(ocppResponse['response'] != 'Accepted')
            strResponse = str(ocppResponse['response'])

        if bSub == False:
            return bResponse
        else:
            return bResponse, strResponse

class LogInfo:
    def __init__(self):
        self.strDateTime  = ''
        self.strEVSEEvent = ''
        self.strCommand   = ''
        self.strTarget    = ''
        self.connector1   = ConnectorInfo()
        self.connector2   = ConnectorInfo()
        self.ocppMessage  = OCPPMessageInfo()

    def clear(self):
        self.strDateTime  = ''
        self.strEVSEEvent = ''
        self.strCommand   = ''
        self.strTarget    = ''
        self.connector1.clear()
        self.connector2.clear()
        self.ocppMessage.clear()

class ReasonInfo:
    def __init__(self):
        self.strCommand        = ''
        self.strPreviousStatus = ''
        self.strCurrentStatus  = ''
        self.strSoC            = ''
        self.strCodeFinish     = ''
        self.strCodeError      = ''
        self.strCodeErrorPLC   = ''
        self.strSequenceName1  = ''
        self.strSequenceName2  = ''

    def clear(self):
        self.strCommand        = ''
        self.strPreviousStatus = ''
        self.strCurrentStatus  = ''
        self.strSoC            = ''
        self.strCodeFinish     = ''
        self.strCodeError      = ''
        self.strCodeErrorPLC   = ''
        self.strSequenceName1  = ''
        self.strSequenceName2  = ''

    def getCodeFinish(self):
        return self.strCodeFinish if len(self.strCodeFinish) > 0 and self.strCodeFinish != '0' else ''

    def getCodeError(self):
        return self.strCodeError if len(self.strCodeError) > 0 and self.strCodeError != '0' else ''

    def getCodeErrorPLC(self):
        return self.strCodeErrorPLC if len(self.strCodeErrorPLC) > 0 and self.strCodeErrorPLC != '0' else ''

    def getDetailCode(self):
        strDetailFinish   = finishToStr(self.strCodeFinish)
        strDetailError    = errorToStr(self.strCodeError)
        strDetailErrorPLC = errorPLCToStr(self.strCodeErrorPLC)

        strDetailCode = ''
        if len(strDetailFinish) > 0:
            strDetailCode += strDetailFinish
        if len(strDetailError) > 0:
            strDetailCode += strDetailError if len(strDetailCode) <= 0 else f' / {strDetailError}'
        if len(strDetailErrorPLC) > 0:
            strDetailCode += strDetailErrorPLC if len(strDetailCode) <= 0 else f' / {strDetailErrorPLC}'

        return strDetailCode

    def getSequenceName1(self):
        if self.strCurrentStatus != 'finish' and self.strCurrentStatus != 'thankYou' and self.strCurrentStatus != 'fault':
            return ''
        return self.strSequenceName1 if '1' in list(self.strSequenceName1[::-1]) else ''

    def getSequenceName2(self):
        if self.strCurrentStatus != 'finish' and self.strCurrentStatus != 'thankYou' and self.strCurrentStatus != 'fault':
            return ''
        return self.strSequenceName2 if '1' in list(self.strSequenceName1[::-1]) else ''

    def getSequenceDetail(self):
        if len(self.strSequenceName1) <= 0 or len(self.strSequenceName2) <= 0:
            return [''] * 32

        strSequenceDetail = self.strSequenceName2 + self.strSequenceName1
        arrSequenceDetail = list(strSequenceDetail[::-1])
        return arrSequenceDetail

class SummaryInfo:
    def __init__(self):
        self.strCharger       = ''
        self.strConnectorId   = ''
        self.strCard          = ''
        self.strExtraData     = ''
        self.strStartDate     = ''
        self.strChargingDate  = ''
        self.strEndDate       = ''
        self.strStartkWh      = ''
        self.strEndkWh        = ''
        self.strTransactionId = ''
        self.strStopReason    = ''
        self.strSubReason     = ''
        self.strTarget        = ''
        self.strCellColor     = ''
        self.reasonInfo       = ReasonInfo()

    def clear(self):
        self.strCharger       = ''
        self.strConnectorId   = ''
        self.strCard          = ''
        self.strExtraData     = ''
        self.strStartDate     = ''
        self.strChargingDate  = ''
        self.strEndDate       = ''
        self.strStartkWh      = ''
        self.strEndkWh        = ''
        self.strTransactionId = ''
        self.strStopReason    = ''
        self.strSubReason     = ''
        self.strTarget        = ''
        self.strCellColor     = ''
        self.reasonInfo.clear()

    def getChargingTime(self):
        strChargingTime = 0
        if len(self.strChargingDate) > 0 and len(self.strEndDate) > 0:
            strChargingTime = calculateTime(self.strEndDate, self.strChargingDate)
        return strChargingTime

    def getTotalTime(self):
        strTotalTime = 0
        if len(self.strStartDate) > 0 and len(self.strEndDate) > 0:
            strTotalTime = calculateTime(self.strEndDate, self.strStartDate)
        return strTotalTime

    def getTotalkWh(self):
        strTotalkWh = 0
        if len(self.strStartkWh) > 0 and len(self.strEndkWh) > 0:
            strTotalkWh = float(self.strEndkWh) - float(self.strStartkWh)
        return f'{strTotalkWh / 1000:.3f}' if strTotalkWh > 0 else ''

    def getHyperLink(self):
        return f'#\'{self.strCharger}\'!{self.strTarget}' if len(self.strTarget) > 0 else None

class AnalyzeInfo:
    def __init__(self):
        self.matched       = 0
        self.connectorId   = 0
        self.ocppUUIDs     = {}
        self.logs          = []
        self.rawLogs       = []
        self.prevLog       = []
        self.prevEmergency = ''

    def clear(self):
        self.matched       = 0
        self.connectorId   = 0
        self.ocppUUIDs     = {}
        self.logs          = []
        self.rawLogs       = []
        self.prevLog       = []
        self.prevEmergency = ''

class Analyzed:
    _instance = None
    _lock = threading.Lock()
    _infos : Dict[str, AnalyzeInfo] = {}

    def __new__(cls):
        if cls._instance == None:
            with cls._lock:
                if cls._instance == None:
                    cls._instance = super(Analyzed, cls).__new__(cls)
        return cls._instance

    def get(self, key):
        with self._lock:
            return self._infos.get(key)

    def get_all(self):
        with self._lock:
            return dict(self._infos)

    def getLogs(self):
        with self._lock:
            return {src: getattr(infos, "logs") for src, infos in self._infos.items()}.items()

    def set(self, key, value):
        if type(value) == AnalyzeInfo:
            with self._lock:
                self._infos[key] = value

    def setLogs(self, key, index, value):
        if type(value) == LogInfo:
            with self._lock:
                self._infos[key].logs[index] = value

    def append(self, key, log):
        with self._lock:
            infos = self._infos.get(key)
            if infos == None:
                self._infos[key] = AnalyzeInfo()
                infos = self._infos.get(key)
            if type(log) == LogInfo:
                infos.logs.append(copy.deepcopy(log))
            elif type(log) == str:
                infos.rawLogs.append(copy.deepcopy(log))

    def sizeLogs(self, key):
        with self._lock:
            infos = self._infos.get(key)
            if infos == None:
                return 0

            return len(infos.logs)

    def remove(self, target):
        with self._lock:
            if target in self._infos:
                del(self._infos[target])

    def clear(self):
        with self._lock:
            self._infos = {}

    def size(self):
        with self._lock:
            return len(self._infos)

# 대상 파일 확인
def getLogFiles():
    loadSiteList()

    # 대상 파일 또는 경로 지정
    if Config().logPath != None:
        Config().logPath = os.path.expanduser(Config().logPath)

        if os.path.exists(Config().logPath) == False:
            print(f'[{now()}] [SCARCommon] getLogFiles - not exist [{Config().logPath}]')
            Config().logPath = None

        # 대상 경로
        elif os.path.isdir(Config().logPath) == False:
            print(f'[{now()}] [SCARCommon] getLogFiles - not a directory [{Config().logPath}]')
            Config().logPath = None

    if Config().logPath == None:
        return 0

    srcPaths = {}
    for path, folders, files in os.walk(Config().logPath):
        if len(folders) > 0:
            for folder in folders:
                srcPaths[folder] = os.path.join(path, folder)
        else:
            srcPaths[''] = Config().logPath

        break

    # info(f'SrcPath - [{srcPaths}]')


    for srcName, srcPath in srcPaths.items():
        for path, folder, files in os.walk(srcPath):
            for file in files:
                if file.find(Config().logName) >= 0 and file.find(Config().zipExt) >= 0:
                    full_path = os.path.join(path, file)

                    #  zip 파일 유효성 검사
                    if not zipfile.is_zipfile(full_path):
                        print(f"⚠️ 유효하지 않은 zip 파일: {full_path}")
                        continue
                    try:
                        with zipfile.ZipFile(full_path) as zip_file:
                            zip_file.extractall(path=path)
                            print(f" 압축 해제 완료: {file}")
                    except zipfile.BadZipFile:
                        print(f" 압축 해제 실패 (BadZipFile): {file}")
                        continue
    LogList().fileCount = 0
    for srcName, srcPath in srcPaths.items():
        #info(f'SrcPath - [{srcName}] [{srcPath}]')

        for path, folder, files in os.walk(srcPath):
            if not os.path.isfile(full_path):
                continue
            srcFiles = []
            dstFiles = ''

            for file in files:
                print(f"[DEBUG] 검사 중: {full_path}")

                if file.find(Config().dstFile) < 0:
                    print(f"  ✅ dstFile 조건 통과: {file}")
                else:
                    print(f"  ❌ dstFile 포함됨: {file}")

                if file.find(Config().logName) >= 0:
                    print(f"  ✅ logName 포함: {file}")
                else:
                    print(f"  ❌ logName 없음: {file}")

                if file.find(Config().logExt) >= 0:
                    print(f"  ✅ logExt 포함: {file}")
                else:
                    print(f"  ❌ logExt 없음: {file}")
                if file.find(Config().dstFile) < 0 and file.find(Config().logName) >= 0 and file.find(Config().logExt) >= 0:
                    srcFiles.append(os.path.join(path, file))
                    print(f'[{now()}] [SCARCommon] Src #{len(srcFiles)} [{os.path.join(path, file)}]')

                    if len(srcName) <= 0:
                        fileNames = parse('{dateTime}_{chargerId}_{etc}', file)
                        if fileNames != None:
                            if fileNames['chargerId'] in Config().siteList:
                                srcName = Config().siteList[fileNames['chargerId']]
                            else:
                                srcName = fileNames['chargerId']
                        else:
                            srcName = 'analyze'

            if len(srcFiles) > 0:
                dstFileName = f'{srcName}{Config().dstFile}'
                dstFiles = os.path.join(Config().logPath, dstFileName)
                print(f'[{now()}] [SCARCommon] Dst #{len(LogList().dstDicts) + 1} [{os.path.join(Config().logPath, dstFileName)}]')

                LogList().fileCount += len(srcFiles)
                if srcName in LogList().srcDicts:
                    tempFiles = LogList().srcDicts[srcName]
                    srcFiles += tempFiles
                LogList().srcDicts[srcName] = sorted(srcFiles)
                LogList().dstDicts[srcName] = dstFiles

    print(f'[{now()}] [SCARCommon] Target - Src [{srcName}] File [{LogList().fileCount}]')

def loadSiteList():
    workbook = None
    worksheet = None

    if os.path.exists(Config().siteFile) == True:
        workbook = load_workbook(Config().siteFile)

        worksheet = workbook['Site List']
        if worksheet != None:
            rowIndex = 2
            while True:
                if worksheet.cell(rowIndex, 1).value != None and worksheet.cell(rowIndex, 14).value != None:
                    Config().siteList[worksheet.cell(rowIndex, 14).value] = worksheet.cell(rowIndex, 1).value
                    rowIndex += 1
                else:
                    break

def getConfig():
    try:
        configPath = os.path.join(os.getcwd(), 'src/pyScript/SCARInfo.yaml')

        with open(configPath) as f:
            config = yaml.safe_load(f)

        Config().logPath    = config['LOG_PATH']
        Config().logName    = config['LOG_NAME']
        Config().logExt     = config['LOG_EXT']
        Config().zipExt     = config['ZIP_EXT']
        Config().dstFile    = config['DST_FILE']
        Config().siteFile   = config['SITE_FILE']
        Config().detailFile = config['DETAIL_FILE']

    except Exception as e:
        print(f'[{now()}] [SCARCommon] getConfig - Exception [{e}]')

def setLogPath(path):
    try:
        configPath = os.path.join(os.getcwd(), 'src/pyScript/SCARInfo.yaml')

        with open(configPath) as f:
            config = yaml.safe_load(f)

        config['LOG_PATH'] = path

        with open(configPath, 'w') as f:
            yaml.dump(config, f)

    except Exception as e:
        print(f'[{now()}] [SCARCommon] setConfig - Exception [{e}]')

# JSON 처리
def parseJson(log, res=False):
    try:
        log = log.replace('\\\\"', '"')
        log = log.replace('\\"', '"')
        log = log.replace('"[', '[')
        log = log.replace(']"', ']')
        log = log.replace('"{', '{')
        log = log.replace('}"', '}')

        if isJson(log) == False:
            # info(f'parseJson - incomplete json')
            return '', dict()

        jArray  = json.loads(log)
        UUID    = jArray[1]
        if res == False:
            jObject = jArray[3]
        else:
            jObject = jArray[2]

        return UUID, jObject

    except Exception as e:
        print(f'[{now()}] [SCARCommon] parseJson - Exception [{e}]')
        return '', dict()

def parseTransactionReq(json):
    ocppInfo = OCPPMessageInfo()

    eventType = json.get('eventType')
    if eventType != None:
        ocppInfo.strEventType = eventType

    evse = json.get('evse')
    if evse != None:
        connectorId = evse.get('connectorId')
        if connectorId != None:
            ocppInfo.strConnectorId = str(connectorId)

    meterValues = json.get('meterValue')
    if meterValues != None:
        for meterValue in meterValues:
            sampledValues = meterValue.get('sampledValue')
            if sampledValues != None:
                for sampledValue in sampledValues:
                    measurand = sampledValue.get('measurand')
                    value = sampledValue.get('value')
                    if measurand == 'Energy.Active.Import.Register' and value != None:
                        ocppInfo.strMeterValue = str(value)
                    elif measurand == 'SoC' and value != None and int(value) != 0:
                        ocppInfo.strSoC = str(value)

            if len(ocppInfo.strMeterValue) > 0 and len(ocppInfo.strSoC) > 0:
                break

    idToken = json.get('idToken')
    if idToken != None:
        subIdToken = idToken.get('idToken')
        if subIdToken != None:
            ocppInfo.strIdTag = subIdToken

    offline = json.get('offline')
    if offline != None:
        ocppInfo.bOffline = bool(offline)

    timestamp = json.get('timestamp')
    if timestamp != None:
        ocppInfo.strTimestamp = timestamp

    transactionInfo = json.get('transactionInfo')
    if transactionInfo != None:
        chargingState = transactionInfo.get('chargingState')
        if chargingState != None:
            ocppInfo.strChargingState = chargingState
        stoppedReason = transactionInfo.get('stoppedReason')
        if stoppedReason != None:
            ocppInfo.strStoppedReason = stoppedReason
        transactionId = transactionInfo.get('transactionId')
        if transactionId != None:
            ocppInfo.strTransactionId = transactionId

    triggerReason = json.get('triggerReason')
    if triggerReason != None:
            ocppInfo.strTriggerReason = triggerReason

    return ocppInfo

def parseTransactionRes(json):
    ocppInfo = OCPPMessageInfo()

    idTokenInfo = json.get('idTokenInfo')
    if idTokenInfo != None:
        status = idTokenInfo.get('status')
        if status != None:
            ocppInfo.strResponse = status

    return ocppInfo

def parseRequestStartReq(json):
    ocppInfo = OCPPMessageInfo()

    idToken = json.get('idToken')
    if idToken != None:
        subIdToken = idToken.get('idToken')
        if subIdToken != None:
            ocppInfo.strIdTag = subIdToken

    return ocppInfo

def parseRequestStartRes(json):
    ocppInfo = OCPPMessageInfo()

    status = json.get('status')
    if status != None:
        ocppInfo.strResponse = status

    return ocppInfo

def parseRequestStopReq(json):
    ocppInfo = OCPPMessageInfo()

    transactionId = json.get('transactionId')
    if transactionId != None:
        ocppInfo.strTransactionId = transactionId

    return ocppInfo

def parseRequestStopRes(json):
    ocppInfo = OCPPMessageInfo()

    status = json.get('status')
    if status != None:
        ocppInfo.strResponse = status

    return ocppInfo

def parseAuthorizeReq(json):
    ocppInfo = OCPPMessageInfo()

    idToken = json.get('idToken')
    if idToken != None:
        subIdToken = idToken.get('idToken')
        if subIdToken != None:
            ocppInfo.strIdTag = subIdToken

    return ocppInfo

def parseAuthorizeRes(json):
    ocppInfo = OCPPMessageInfo()

    idTokenInfo = json.get('idTokenInfo')
    if idTokenInfo != None:
        status = idTokenInfo.get('status')
        if status != None:
            ocppInfo.strResponse = status

    return ocppInfo

def parseNotificationReq(json):
    ocppInfo = OCPPMessageInfo()

    connectorId = json.get('connectorId')
    if connectorId != None:
        ocppInfo.strConnectorId = str(connectorId)

    connectorStatus = json.get('connectorStatus')
    if connectorStatus != None:
        ocppInfo.strEventType = connectorStatus

    return ocppInfo

def parseNotifyReq(json):
    ocppInfo = OCPPMessageInfo()

    eventDatas = json.get('eventData')
    if eventDatas != None:
        for eventData in eventDatas:
            component = eventData.get('component')
            if component != None:
                evse = component.get('evse')
                if evse != None:
                    connectorId = evse.get('connectorId')
                    if connectorId != None:
                        ocppInfo.strConnectorId = str(connectorId)

            techInfo = eventData.get('techInfo')
            if techInfo != None:
                ocppInfo.strTriggerReason = techInfo

    return ocppInfo

def isJson(object):
    try:
        jObject = json.loads(object)    # { } 가 포함된 string이 invalid json 인 경우 Exception
        iterator = iter(jObject)        # { } 가 없는 경우는 string의 경우 Exception

    except Exception as e:
        # info(f'Exception [{e}]')
        return False

    return True

def now():
    now = datetime.now()
    return now.isoformat()

def convertTime(strTime: str, nOffset: int = 9) -> str:
    try:
        utcTime = datetime.fromisoformat(strTime.replace('Z', '+00:00'))

        zOffset = timezone(timedelta(hours=nOffset))

        # Convert to Asia/Seoul timezone
        convertedTime = utcTime.astimezone(zOffset)
        return convertedTime.strftime('%Y-%m-%d %H:%M:%S')

    except Exception as e:
        raise ValueError(f'convertTime - Exception [{e}]')

def calculateTime(strTime1, strTime2):
    time1 = datetime.fromisoformat(strTime1.replace('Z', '+00:00'))
    time2 = datetime.fromisoformat(strTime2.replace('Z', '+00:00'))

    return int((time1 - time2).total_seconds())

def secsToStr(seconds):
    hh = seconds // 3600
    mm = (seconds % 3600) // 60
    ss = seconds % 60

    return f'{hh:02}:{mm:02}:{ss:02}' if seconds > 0 else ''

def finishToStr(finish):
    dictFinishCode = {  '1' : 'FINISH_CODE_USER_STOP',
                        '2' : 'FINISH_CODE_SESSION_STOP',
                        '3' : 'FINISH_CODE_S3_STOP',
                        '4' : 'FINISH_CODE_PAYMENT_TIMEOUT',
                        '5' : 'FINISH_CODE_EV_STOP',
                        '6' : 'FINISH_CODE_READY_TO_CHARGE_OFF' }
    strFinishCode = ''
    if finish in dictFinishCode:
        strFinishCode = dictFinishCode[finish]
    return strFinishCode

def errorToStr(error):
    dictErrorCode = {   '3'    : 'ERR_CODE_PC_MC_CONTACT_ERR',
                        '7'    : 'ERR_CODE_PWR_MD_AC_UNDER',
                        '8'    : 'ERR_CODE_PWR_MD_AC_OVER',
                        '10'   : 'ERR_CODE_OVER_VOLTAGE',
                        '11'   : 'ERR_CODE_OVER_CURRENT',
                        '13'   : 'ERR_CODE_EMG_SW_PUSH',
                        '14'   : 'ERR_CODE_IMD_FAULT',
                        '41'   : 'ERR_CODE_PRECHARGE_VOLTAGE_NOT_ENOUGH',
                        '42'   : 'ERR_CODE_SECC_REPORTINIT_RES_TIMEOUT',
                        '44'   : 'ERR_CODE_SECC_PRECOMMUNICATION_TIMEOUT',
                        '45'   : 'ERR_CODE_SECC_CHARGE_PARAMETER_TIMEOUT',
                        '46'   : 'ERR_CODE_SECC_CABLE_CHECK_TIMEOUT',
                        '47'   : 'ERR_CODE_SECC_PRECHARGE_TIMEOUT',
                        '48'   : 'ERR_CODE_SECC_CURRENT_DEMAND_TIMEOUT',
                        '49'   : 'ERR_CODE_COUPLER_TEMP_HIGH',
                        '53'   : 'ERR_CODE_CONTRACT_TIMEOUT',
                        '54'   : 'ERR_CODE_CP_LEVEL_12V',
                        '56'   : 'ERR_CODE_CP_LEVEL_PE',
                        '74'   : 'ERR_CODE_PP_LEVEL',
                        '97'   : 'ERR_CODE_MCU_HARD_FAULT',
                        '121'  : 'ERR_CODE_HMI_COMMUNICATION',
                        '218'  : 'ERR_CODE_IMD_COMMUNICATION',
                        '242'  : 'ERR_CODE_DP_DOOR',
                        '243'  : 'ERR_CODE_CP_DOOR',
                        '244'  : 'ERR_CODE_BANK_ERROR',
                        '245'  : 'ERR_CODE_IMD_STATE',
                        '260'  : 'ERR_CODE_N_TEP_COMMUNICATION',
                        '261'  : 'ERR_CODE_PC_COMMUNICATION',
                        '262'  : 'ERR_CODE_PC_RELAYBOX_COMM_ERROR',
                        '263'  : 'ERR_CODE_PC_ASSIGN_FAIL_ERROR',
                        '264'  : 'ERR_CODE_PC_RELAYBOX_PWR_MD_FAULT',
                        '265'  : 'ERR_CODE_PC_DC_SHORT',
                        '266'  : 'ERR_CODE_PWR_MD',
                        '270'  : 'ERR_CODE_DP_ID_COLISION',
                        '271'  : 'ERR_CODE_SECC_ERRCODE',
                        '272'  : 'ERR_CODE_SECC_READY_TO_CHARGE',
                        '281'  : 'ERR_CODE_SECC_REQ_OVER_VOLTAGE',
                        '282'  : 'ERR_CODE_SECC_OUT_MAX_PWR_OVER',
                        '283'  : 'ERR_CODE_SECC_OUT_MAX_V_OVER',
                        '284'  : 'ERR_CODE_SECC_OUT_MAX_A_OVER',
                        '285'  : 'ERR_CODE_CP_FG_SHORT',
                        '286'  : 'ERR_CODE_CABLECHECK_VOLTAGE_FLOAT',
                        '287'  : 'ERR_CODE_IMD_VOLTAGE_ERR',
                        '288'  : 'ERR_CODE_ADC_VOLTAGE_ERR',
                        '289'  : 'ERR_CODE_ADC_CURRENT_ERR',
                        '300'  : 'ERR_CODE_WARING_EV_MAX_VOLTAGE_OVER',
                        '301'  : 'ERR_CODE_PB_ERR_ID_COLLISOIN',
                        '302'  : 'ERR_CODE_PB_ERR_CAN1_COMM',
                        '303'  : 'ERR_CODE_PB_ERR_CAN2_COMM',
                        '304'  : 'ERR_CODE_PB_ERR_PS_LOGIC_COLLISION',
                        '305'  : 'ERR_CODE_CH_ERR_ASSING_FAIL',
                        '306'  : 'ERR_CODE_CHARGING_VOLTAGE_OUT_OF_RANGE',
                        '307'  : 'ERR_CODE_EV_ERROR_Reserved_A',
                        '308'  : 'ERR_CODE_EV_ERROR_Reserved_B',
                        '309'  : 'ERR_CODE_CH_ERR_ALL_RLYBOX_ERR',
                        '310'  : 'ERR_CODE_CHARGING_SYSTEM_INCOMPATIBILITY',
                        '311'  : 'ERR_CODE_EV_NO_DATA',
                        '1001' : 'ERR_CODE_OVER_HMI_PWR',
                        '1002' : 'ERR_CODE_OVER_CPL_CUR',
                        '1003' : 'ERR_CODE_OVER_CURRENT',
                        '1101' : 'ERR_CODE_ELCB_SHUTDOWN',
                        '1102' : 'ERR_CODE_FUSE_DISCONNECT',
                        '1103' : 'ERR_CODE_AC_SPD_SHUTDOWN',
                        '1104' : 'ERR_CODE_PWR_MD_AC_UNDER',
                        '1105' : 'ERR_CODE_PWR_MD_AC_OVER',
                        '1106' : 'ERR_CODE_PWR_MD_CUR_OVER',
                        '1107' : 'ERR_CODE_PWR_MD_AC_CUR_OVER_L1',
                        '1108' : 'ERR_CODE_PWR_MD_AC_CUR_OVER_L2',
                        '1109' : 'ERR_CODE_PWR_MD_AC_CUR_OVER_L3',
                        '1201' : 'ERR_CODE_NON_PWR_MD_IN_GRUP',
                        '2001' : 'ERR_CODE_PWR_MD_01_DC_SHORT',
                        '2002' : 'ERR_CODE_PWR_MD_02_DC_SHORT',
                        '2003' : 'ERR_CODE_PWR_MD_03_DC_SHORT',
                        '2004' : 'ERR_CODE_PWR_MD_04_DC_SHORT',
                        '2005' : 'ERR_CODE_PWR_MD_05_DC_SHORT',
                        '2006' : 'ERR_CODE_PWR_MD_06_DC_SHORT',
                        '2007' : 'ERR_CODE_PWR_MD_07_DC_SHORT',
                        '2008' : 'ERR_CODE_PWR_MD_08_DC_SHORT',
                        '2101' : 'ERR_CODE_PWR_MD_01_SHORT_CIRCUIT',
                        '2102' : 'ERR_CODE_PWR_MD_02_SHORT_CIRCUIT',
                        '2103' : 'ERR_CODE_PWR_MD_03_SHORT_CIRCUIT',
                        '2104' : 'ERR_CODE_PWR_MD_04_SHORT_CIRCUIT',
                        '2105' : 'ERR_CODE_PWR_MD_05_SHORT_CIRCUIT',
                        '2106' : 'ERR_CODE_PWR_MD_06_SHORT_CIRCUIT',
                        '2107' : 'ERR_CODE_PWR_MD_07_SHORT_CIRCUIT',
                        '2108' : 'ERR_CODE_PWR_MD_08_SHORT_CIRCUIT',
                        '2201' : 'ERR_CODE_PWR_MD_01_BLEED_NOT_WORK',
                        '2202' : 'ERR_CODE_PWR_MD_02_BLEED_NOT_WORK',
                        '2203' : 'ERR_CODE_PWR_MD_03_BLEED_NOT_WORK',
                        '2204' : 'ERR_CODE_PWR_MD_04_BLEED_NOT_WORK',
                        '2205' : 'ERR_CODE_PWR_MD_05_BLEED_NOT_WORK',
                        '2206' : 'ERR_CODE_PWR_MD_06_BLEED_NOT_WORK',
                        '2207' : 'ERR_CODE_PWR_MD_07_BLEED_NOT_WORK',
                        '2208' : 'ERR_CODE_PWR_MD_08_BLEED_NOT_WORK',
                        '3001' : 'ERR_CODE_IMD_VOL_DAMAGE',
                        '3101' : 'ERR_CODE_DC_METER_COMM',
                        '3102' : 'ERR_CODE_DC_METER_VOL_DAMAGE',
                        '3103' : 'ERR_CODE_DC_METER_CUR_DAMAGE',
                        '4001' : 'ERR_CODE_WELDING_MC',
                        '4101' : 'ERR_CODE_WELDING_RELAY01',
                        '4102' : 'ERR_CODE_WELDING_RELAY02',
                        '4103' : 'ERR_CODE_WELDING_RELAY03',
                        '4104' : 'ERR_CODE_WELDING_RELAY04',
                        '4105' : 'ERR_CODE_WELDING_RELAY05',
                        '4106' : 'ERR_CODE_WELDING_RELAY06',
                        '4201' : 'ERR_CODE_MALFUNC_MC',
                        '4301' : 'ERR_CODE_MALFUNC_RELAY01',
                        '4302' : 'ERR_CODE_MALFUNC_RELAY02',
                        '4303' : 'ERR_CODE_MALFUNC_RELAY03',
                        '4304' : 'ERR_CODE_MALFUNC_RELAY04',
                        '4305' : 'ERR_CODE_MALFUNC_RELAY05',
                        '4306' : 'ERR_CODE_MALFUNC_RELAY06',
                        '4401' : 'ERR_CODE_TILT_SENSOR',
                        '4404' : 'ERR_CODE_OVER_TEMP',
                        '4410' : 'ERR_CODE_SUBMERSION_TEMP',
                        '9001' : 'ERR_CODE_DUAL_1CHSINGLE_POWERSHARING_MALFUNC',
                        '9002' : 'ERR_CODE_1CHSINGLE_DUAL_POWERSHARING_MALFUNC',
                        '9003' : 'ERR_CODE_DUAL_2CHSINGLE_POWERSHARING_MALFUNC',
                        '9004' : 'ERR_CODE_2CHSINGLE_DUAL_POWERSHARING_MALFUNC',
                        '9011' : 'ERR_CODE_2CH_RELAY_MALFUNC',
                        '9012' : 'ERR_CODE_1CH_RELAY_MALFUNC',
                        '9013' : 'ERR_CODE_BRIDGE_RELAY_MALFUNC',
                        '9100' : 'ERR_CODE_PWR_MD_NOT_AVAILABLE_MALFUNC'    }

    strErrorCode = ''
    if error in dictErrorCode:
        strErrorCode = dictErrorCode[error]
    return strErrorCode

def errorPLCToStr(plc):
    dictPLCCode = { '17'  : 'SECC_INIT_ERROR_IFADDR',
                    '18'  : 'SECC_INIT_ERROR_THREAD',
                    '19'  : 'SECC_INIT_ERROR_OPENCHANNEL',
                    '20'  : 'SECC_INIT_ERROR_KEY',
                    '33'  : 'SECC_SLAC_ERROR_TIMER_INIT',
                    '34'  : 'SECC_SLAC_ERROR_TIMER_TIMEOUT',
                    '36'  : 'SECC_SLAC_ERROR_PARAM_TIMEOUT',
                    '37'  : 'SECC_SLAC_ERROR_PARAM_SOCKET',
                    '38'  : 'SECC_SLAC_ERROR_START_ATTEN_CHAR_TIMEOUT',
                    '39'  : 'SECC_SLAC_ERROR_MNBC_SOUND_TIMEOUT',
                    '40'  : 'SECC_SLAC_ERROR_ATTEN_CHAR_TIMEOUT',
                    '41'  : 'SECC_SLAC_ERROR_ATTEN_CHAR_SOCKET',
                    '47'  : 'SECC_SLAC_ERROR_MATCH_TIMEOUT',
                    '48'  : 'SECC_SLAC_ERROR_MATCH_SOCKET',
                    '51'  : 'SECC_SLAC_ERROR_MATCH_TIMEOUT2',
                    '65'  : 'SECC_SDP_ERROR_INIT_SOCKET',
                    '66'  : 'SECC_SDP_ERROR_INIT_SOCKOPT1',
                    '67'  : 'SECC_SDP_ERROR_INIT_SOCKOPT2',
                    '68'  : 'SECC_SDP_ERROR_INIT_BIND',
                    '69'  : 'SECC_SDP_ERROR_THREAD_SOCKET1',
                    '70'  : 'SECC_SDP_ERROR_THREAD_SOCKET2',
                    '71'  : 'SECC_SDP_ERROR_TIMEOUT',
                    '80'  : 'SECC_DIN_ERROR_GENERAL',
                    '81'  : 'SECC_DIN_ERROR_INIT_SOCKET',
                    '82'  : 'SECC_DIN_ERROR_INIT_SOCKOPT',
                    '83'  : 'SECC_DIN_ERROR_INIT_BIND',
                    '84'  : 'SECC_DIN_ERROR_INIT_LISTEN',
                    '85'  : 'SECC_DIN_ERROR_INIT_SELECT',
                    '86'  : 'SECC_DIN_ERROR_INIT_ACCEPT',
                    '87'  : 'SECC_DIN_ERROR_TIMEOUT',
                    '88'  : 'SECC_DIN_ERROR_V2GTP_HEADER',
                    '89'  : 'SECC_DIN_ERROR_V2GTP_HEADER_LEN',
                    '90'  : 'SECC_DIN_ERROR_DECODE_EXI',
                    '91'  : 'SECC_DIN_ERROR_CREATE_RESPONSE',
                    '92'  : 'SECC_DIN_ERROR_ENCODE_EXI',
                    '93'  : 'SECC_DIN_ERROR_V2GTP_HEADER_WRITE',
                    '94'  : 'SECC_DIN_ERROR_SOCKET_EXCEPTION',
                    '95'  : 'SECC_DIN_ERROR_SOCKET_SEND',
                    '96'  : 'SECC_DIN_ERROR_NO_PROTOCOL',
                    '128' : 'SECC_TLS_ERROR_HELLO_REQUEST',
                    '129' : 'SECC_TLS_ERROR_CLIENT_HELLO',
                    '130' : 'SECC_TLS_ERROR_SERVER_HELLO',
                    '131' : 'SECC_TLS_ERROR_SERVER_CERTIFICATE',
                    '132' : 'SECC_TLS_ERROR_SERVER_KEY_EXCHANGE',
                    '133' : 'SECC_TLS_ERROR_CERTIFICATE_REQUEST',
                    '134' : 'SECC_TLS_ERROR_SERVER_HELLO_DONE',
                    '135' : 'SECC_TLS_ERROR_CLIENT_CERTIFICATE',
                    '136' : 'SECC_TLS_ERROR_CLIENT_KEY_EXCHANGE',
                    '137' : 'SECC_TLS_ERROR_CERTIFICATE_VERIFY',
                    '138' : 'SECC_TLS_ERROR_CLIENT_CHANGE_CIPHER_SPEC',
                    '139' : 'SECC_TLS_ERROR_CLIENT_FINISHED',
                    '140' : 'SECC_TLS_ERROR_SERVER_CHANGE_CIPHER_SPEC',
                    '141' : 'SECC_TLS_ERROR_SERVER_FINISHED',
                    '142' : 'SECC_TLS_ERROR_FLUSH_BUFFERS',
                    '143' : 'SECC_TLS_ERROR_HANDSHAKE_WRAPUP',
                    '144' : 'SECC_TLS_ERROR_HANDSHAKE_OVER',
                    '145' : 'SECC_TLS_ERROR_SERVER_NEW_SESSION_TICKET',
                    '146' : 'SECC_TLS_ERROR_SERVER_HELLO_VERIFY_REQUEST_SENT',
                    '147' : 'SECC_TLS_ERROR_ALERT_FATAL'    }

    strPLCCode = ''
    if plc in dictPLCCode:
        strPLCCode = dictPLCCode[plc]
    return strPLCCode
